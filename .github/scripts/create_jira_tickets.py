#!/usr/bin/env python3
"""
create_jira_tickets.py
──────────────────────
Workflow 1 / Step 3 — Jira Manager

Reads the sorted Excel produced by fetch_dependabot_alerts.py,
groups alerts by service, and for each service:
  - searches Jira for an existing open GHAS ticket (dedup check)
  - creates a new ticket if none found
  - updates the existing ticket if one is found
  - writes the Jira key and status back into the Excel file

Requirements:
    pip install requests openpyxl python-dotenv

Environment Variables (read from .env at repo root):
    JIRA_URL          — https://your-org.atlassian.net
    JIRA_EMAIL        — Atlassian account email
    JIRA_API_TOKEN    — Atlassian API token
    JIRA_PROJECT_KEY  — Jira project key (e.g. SCRUM)

Usage:
    python create_jira_tickets.py                           # picks latest Excel in cwd
    python create_jira_tickets.py dependabot_alerts_X.xlsx  # specific file
"""

import os
import sys
import glob as glob_module
import json
import base64
import requests
from datetime import date
from openpyxl import load_workbook
from dotenv import load_dotenv

# Load .env from repo root (two levels up from .github/scripts/)
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", "..", ".env"))

# ── CONFIG ────────────────────────────────────────────────────────────────────
JIRA_URL        = os.getenv("JIRA_URL", "").rstrip("/")
JIRA_EMAIL      = os.getenv("JIRA_EMAIL", "")
JIRA_API_TOKEN  = os.getenv("JIRA_API_TOKEN", "")
PROJECT_KEY     = os.getenv("JIRA_PROJECT_KEY", "SCRUM")
SCAN_DATE       = date.today().isoformat()

SEVERITY_ORDER  = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
PRIORITY_MAP    = {
    "CRITICAL": "Highest",
    "HIGH":     "High",
    "MEDIUM":   "Medium",
    "LOW":      "Low",
}

# ADF status badge colors (renders as colored pill in Jira Cloud UI)
SEVERITY_BADGE_COLOR = {
    "CRITICAL": "red",
    "HIGH":     "yellow",
    "MEDIUM":   "blue",
    "LOW":      "green",
}

EXCEL_COLUMNS = [
    "Service", "Repo", "Alert #", "Severity", "CVE ID", "GHSA ID",
    "Package", "Vulnerable Range", "Safe Version",
    "Manifest", "Scope", "Summary", "Alert URL",
    "Jira Key", "Jira Status",
]
COL = {name: idx for idx, name in enumerate(EXCEL_COLUMNS)}


# ── VALIDATION ────────────────────────────────────────────────────────────────
def validate_config():
    missing = [k for k, v in {
        "JIRA_URL": JIRA_URL, "JIRA_EMAIL": JIRA_EMAIL,
        "JIRA_API_TOKEN": JIRA_API_TOKEN, "JIRA_PROJECT_KEY": PROJECT_KEY,
    }.items() if not v]
    if missing:
        print(f"ERROR: Missing required env vars: {', '.join(missing)}")
        print("       Set them in the .env file at the repo root.")
        sys.exit(1)


# ── JIRA AUTH & HTTP ──────────────────────────────────────────────────────────
def _auth_header() -> dict:
    creds = base64.b64encode(f"{JIRA_EMAIL}:{JIRA_API_TOKEN}".encode()).decode()
    return {
        "Authorization": f"Basic {creds}",
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }


def _get(path: str, params: dict = None) -> dict:
    resp = requests.get(f"{JIRA_URL}{path}", headers=_auth_header(), params=params)
    resp.raise_for_status()
    return resp.json()


def _post(path: str, body: dict) -> dict:
    resp = requests.post(f"{JIRA_URL}{path}", headers=_auth_header(), json=body)
    resp.raise_for_status()
    return resp.json()


def _put(path: str, body: dict):
    resp = requests.put(f"{JIRA_URL}{path}", headers=_auth_header(), json=body)
    resp.raise_for_status()

def _add_comment(issue_key: str, adf_body: dict):
    _post(f"/rest/api/3/issue/{issue_key}/comment", {"body": adf_body})


# ── EXCEL ─────────────────────────────────────────────────────────────────────
def find_excel() -> str:
    if len(sys.argv) > 1:
        path = sys.argv[1]
        if not os.path.exists(path):
            print(f"ERROR: File not found: {path}")
            sys.exit(1)
        return path
    files = sorted(glob_module.glob("dependabot_alerts_*.xlsx"), reverse=True)
    if not files:
        print("ERROR: No dependabot_alerts_*.xlsx found in current directory.")
        print("       Run fetch_dependabot_alerts.py first.")
        sys.exit(1)
    latest = files[0]
    print(f"[EXCEL] Using: {latest}")
    return latest


def read_and_group(excel_path: str) -> dict:
    """Read Alerts sheet and return dict keyed by service name."""
    wb = load_workbook(excel_path)
    ws = wb["Alerts"]
    groups: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        svc = row[COL["Service"]]
        if not svc:
            continue
        alert = {col: row[idx] for col, idx in COL.items()}
        groups.setdefault(str(svc), []).append(alert)
    return groups


def update_excel(excel_path: str, results: dict):
    """Write Jira Key and Jira Status back to every row of each service."""
    wb = load_workbook(excel_path)
    ws = wb["Alerts"]
    col_key    = COL["Jira Key"] + 1    # openpyxl columns are 1-indexed
    col_status = COL["Jira Status"] + 1
    for row in ws.iter_rows(min_row=2):
        svc = row[COL["Service"]].value
        if svc and str(svc) in results:
            row[col_key - 1].value    = results[str(svc)]["key"]
            row[col_status - 1].value = results[str(svc)]["status"]
    wb.save(excel_path)
    print(f"\n[EXCEL] ✅ Saved: {excel_path}")


# ── COUNTS & TITLE ────────────────────────────────────────────────────────────
def severity_counts(alerts: list) -> dict:
    counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
    for a in alerts:
        sev = str(a.get("Severity") or "").upper()
        if sev in counts:
            counts[sev] += 1
    return counts


def build_title(service: str, counts: dict) -> str:
    sev_names = {"CRITICAL": "Critical", "HIGH": "High", "MEDIUM": "Medium", "LOW": "Low"}
    parts = [f"{sev_names[k]}-{v}" for k, v in counts.items() if v > 0]
    return f"Address GHAS vulnerabilities for {service} [{', '.join(parts)}]"


def counts_label(counts: dict) -> str:
    return "  ".join(f"C-{counts['CRITICAL']} | H-{counts['HIGH']} | M-{counts['MEDIUM']} | L-{counts['LOW']}")


def top_severity(counts: dict) -> str:
    for sev in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
        if counts[sev] > 0:
            return sev
    return "LOW"


# ── ADF HELPERS ───────────────────────────────────────────────────────────────
# Atlassian Document Format (ADF) — required for Jira Cloud REST API v3

def _text(t: str) -> dict:
    return {"type": "text", "text": str(t or "")}


def _bold(t: str) -> dict:
    return {"type": "text", "text": str(t or ""), "marks": [{"type": "strong"}]}


def _para(*content) -> dict:
    return {"type": "paragraph", "content": list(content)}


def _heading(level: int, text: str) -> dict:
    return {"type": "heading", "attrs": {"level": level},
            "content": [{"type": "text", "text": text}]}


def _rule() -> dict:
    return {"type": "rule"}


def _th(text: str) -> dict:
    return {"type": "tableHeader", "attrs": {},
            "content": [{"type": "paragraph", "content": [_bold(text)]}]}


def _td(text: str) -> dict:
    return {"type": "tableCell", "attrs": {},
            "content": [{"type": "paragraph", "content": [_text(str(text or ""))]}]}


def _td_badge(sev: str) -> dict:
    """Table cell with a color-coded ADF status badge (no emoji)."""
    return {
        "type": "tableCell", "attrs": {},
        "content": [{
            "type": "paragraph",
            "content": [{
                "type": "status",
                "attrs": {
                    "text":  sev,
                    "color": SEVERITY_BADGE_COLOR.get(sev, "neutral"),
                },
            }],
        }],
    }


def _table_row(cells: list) -> dict:
    return {"type": "tableRow", "content": cells}


def _build_vuln_table(alerts: list) -> dict:
    """Single table with all CVEs — sorted by severity then package name."""
    header = _table_row([
        _th("#"), _th("Severity"), _th("CVE ID"), _th("GHSA ID"),
        _th("Package"), _th("Vulnerable Range"), _th("Safe Version"), _th("Summary"),
    ])
    sorted_alerts = sorted(
        alerts,
        key=lambda a: (
            SEVERITY_ORDER.get(str(a.get("Severity") or "").upper(), 99),
            str(a.get("Package") or ""),
        ),
    )
    rows = [header]
    for i, a in enumerate(sorted_alerts, 1):
        sev = str(a.get("Severity") or "").upper()
        rows.append(_table_row([
            _td(str(i)),
            _td_badge(sev),
            _td(a.get("CVE ID", "N/A")),
            _td(a.get("GHSA ID", "N/A")),
            _td(a.get("Package", "")),
            _td(a.get("Vulnerable Range", "")),
            _td(a.get("Safe Version", "")),
            _td(a.get("Summary", "")),
        ]))
    return {
        "type": "table",
        "attrs": {"isNumberColumnEnabled": False, "layout": "full-width"},
        "content": rows,
    }


def build_adf_description(service: str, alerts: list) -> dict:
    repo   = str(alerts[0].get("Repo", "")) if alerts else ""
    counts = severity_counts(alerts)
    total  = sum(counts.values())

    sev_summary = "   ".join(
        f"{k.capitalize()}: {v}" for k, v in counts.items() if v > 0
    )

    content = [
        _heading(2, f"Dependabot Security Alerts - {service}"),
        _para(_bold("Repo: "),       _text(repo)),
        _para(_bold("Scan Date: "),  _text(SCAN_DATE)),
        _para(_bold("Total Alerts: "), _text(f"{total}  ({sev_summary})")),
        _rule(),
        _heading(3, f"Vulnerability Report - {service}"),
        _build_vuln_table(alerts),
        _rule(),
        _para(_text(
            "Auto-created by GHAS Vulnerability Management - Workflow 1 / Jira Manager"
        )),
    ]
    return {"version": 1, "type": "doc", "content": content}


def build_adf_comment(counts: dict) -> dict:
    total  = sum(counts.values())
    clabel = counts_label(counts)
    return {
        "version": 1, "type": "doc",
        "content": [
            _para(_text(f"Re-scanned on {SCAN_DATE}")),
            _para(_bold("Alert counts updated: "), _text(clabel)),
            _para(_bold("Total alerts: "), _text(str(total))),
            _rule(),
            _para(_text(
                "Automated by GHAS Vulnerability Management - Workflow 1 / Jira Manager"
            )),
        ],
    }


# ── JIRA OPERATIONS ───────────────────────────────────────────────────────────
def jira_search(service: str) -> str | None:
    """Return existing open ticket key for this service, or None."""
    jql = (
        f'project = "{PROJECT_KEY}" '
        f'AND labels = "{service}" '
        f'AND labels = "dependabot" '
        f'AND statusCategory in ("To Do", "In Progress")'
    )
    try:
        # Jira Cloud REST API v3 uses POST /rest/api/3/search/jql
        data = _post("/rest/api/3/search/jql", {
            "jql": jql, "fields": ["summary", "status"], "maxResults": 1
        })
        issues = data.get("issues", [])
        return issues[0]["key"] if issues else None
    except requests.HTTPError as e:
        raise RuntimeError(f"Jira search failed for '{service}': {e}") from e


def create_ticket(service: str, alerts: list) -> str:
    """Create a new Jira ticket and return its key."""
    counts  = severity_counts(alerts)
    title   = build_title(service, counts)
    top_sev = top_severity(counts)
    body = {
        "fields": {
            "project":     {"key": PROJECT_KEY},
            "issuetype":   {"name": "Task"},
            "summary":     title,
            "priority":    {"name": PRIORITY_MAP[top_sev]},
            "labels":      ["GHAS", service, "dependabot", "security"],
            "description": build_adf_description(service, alerts),
        }
    }
    try:
        result = _post("/rest/api/3/issue", body)
        return result["key"]
    except requests.HTTPError as e:
        raise RuntimeError(f"Ticket creation failed for '{service}': {e.response.text}") from e


def update_ticket(ticket_key: str, service: str, alerts: list):
    """Update title + description on an existing ticket and add a comment."""
    counts = severity_counts(alerts)
    title  = build_title(service, counts)
    body   = {
        "fields": {
            "summary":     title,
            "description": build_adf_description(service, alerts),
        }
    }
    try:
        _put(f"/rest/api/3/issue/{ticket_key}", body)
        _add_comment(ticket_key, build_adf_comment(counts))
    except requests.HTTPError as e:
        raise RuntimeError(f"Ticket update failed for '{ticket_key}': {e.response.text}") from e


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  Workflow 1 — Jira Manager")
    print("=" * 60)

    validate_config()

    excel_path = find_excel()
    groups     = read_and_group(excel_path)

    if not groups:
        print("\n[OK] Excel has no alert rows. Nothing to do.")
        sys.exit(0)

    services = sorted(groups.keys())
    print(f"\n[INFO] Services found: {', '.join(services)}")
    print(f"[INFO] Jira project  : {PROJECT_KEY} ({JIRA_URL})\n")

    results: dict[str, dict] = {}    # service → {key, status}
    created, updated, skipped, failed = [], [], [], []

    for svc in services:
        alerts = groups[svc]
        counts = severity_counts(alerts)
        print(f"[JIRA] Processing: {svc}  "
              f"(C-{counts['CRITICAL']} H-{counts['HIGH']} "
              f"M-{counts['MEDIUM']} L-{counts['LOW']})")

        # ── Dedup check ───────────────────────────────────────────────────────
        try:
            existing_key = jira_search(svc)
        except RuntimeError as e:
            print(f"  ✗ Search error: {e}")
            failed.append(svc)
            results[svc] = {"key": "ERROR", "status": "FAILED"}
            continue

        # ── Create or Update ──────────────────────────────────────────────────
        if existing_key:
            print(f"  -> Found existing ticket: {existing_key} - will UPDATE")
            try:
                update_ticket(existing_key, svc, alerts)
                print(f"  ✅ Updated: {existing_key}")
                updated.append({"service": svc, "key": existing_key})
                results[svc] = {"key": existing_key, "status": "UPDATED"}
            except RuntimeError as e:
                print(f"  ✗ Update failed: {e}")
                failed.append(svc)
                results[svc] = {"key": existing_key, "status": "UPDATE_FAILED"}
        else:
            print(f"  -> No existing ticket found - will CREATE")
            try:
                new_key = create_ticket(svc, alerts)
                print(f"  ✅ Created: {new_key}")
                created.append({"service": svc, "key": new_key})
                results[svc] = {"key": new_key, "status": "CREATED"}
            except RuntimeError as e:
                print(f"  ✗ Creation failed: {e}")
                failed.append(svc)
                results[svc] = {"key": "ERROR", "status": "FAILED"}

    # ── Write results back to Excel ───────────────────────────────────────────
    update_excel(excel_path, results)

    # ── Final summary ─────────────────────────────────────────────────────────
    print("\n" + "=" * 62)
    print("  WORKFLOW 1 — JIRA MANAGER COMPLETE")
    print("=" * 62)
    print(f"  Excel report   : {excel_path}")
    print(f"  Services found : {len(services)}")
    total_alerts = sum(len(v) for v in groups.values())
    print(f"  Total alerts   : {total_alerts}")
    print(f"  Jira CREATED   : {len(created)}  → {[x['key'] for x in created]}")
    print(f"  Jira UPDATED   : {len(updated)}  → {[x['key'] for x in updated]}")
    print(f"  Jira FAILED    : {len(failed)}  → {failed}")
    print("=" * 62)

    if created or updated:
        print("\n  Services with tickets (ready for Workflow 2):")
        for svc, info in results.items():
            if info["status"] in ("CREATED", "UPDATED"):
                print(f"    • {svc:<25} → {info['key']}")

    if failed:
        print(f"\n  ⚠️  {len(failed)} service(s) failed — check errors above and retry.")
        sys.exit(1)


if __name__ == "__main__":
    main()
